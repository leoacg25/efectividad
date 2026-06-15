"""
Genera un archivo Excel de ejemplo para probar el Dashboard de Programadores.
Cada hoja representa un programador con tickets en el formato requerido.
"""
import subprocess, sys

# Instalar openpyxl si no está disponible
try:
    import openpyxl
except ImportError:
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
    import openpyxl

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import random, os

# Datos de ejemplo
PROGRAMMERS = {
    "Carlos Mendez": [
        ("TK-001", "Corrección de bug en módulo de pagos", "PaymentGateway", "Prioridad alta"),
        ("TK-002", "Implementar endpoint REST de usuarios", "UserService", "Requiere autenticación JWT"),
        ("TK-003", "Optimizar consultas SQL en reportes", "ReportsModule", "Ver análisis de EXPLAIN"),
        ("TK-004", "Migrar componentes a React 18", "FrontendApp", ""),
        ("TK-005", "Setup CI/CD pipeline", "DevOps", "GitHub Actions"),
        ("TK-006", "Documentar API con Swagger", "UserService", ""),
        ("TK-007", "Fix memory leak en worker service", "BackgroundJobs", "Detectado en producción"),
        ("TK-008", "Implementar caché Redis", "PaymentGateway", "TTL de 5 minutos"),
    ],
    "Ana García": [
        ("TK-009", "Diseñar UI del dashboard de analytics", "Analytics", "Wireframes en Figma"),
        ("TK-010", "Integrar Chart.js en módulo de reportes", "Analytics", ""),
        ("TK-011", "Responsive design para mobile", "FrontendApp", "Breakpoints: 768px, 480px"),
        ("TK-012", "Accesibilidad WCAG 2.1", "FrontendApp", "Nivel AA requerido"),
        ("TK-013", "Optimizar imágenes y assets", "FrontendApp", "WebP + lazy loading"),
        ("TK-014", "Implementar dark mode", "FrontendApp", "CSS variables"),
    ],
    "Luis Torres": [
        ("TK-015", "Configurar base de datos PostgreSQL", "Infrastructure", "Replicación master-slave"),
        ("TK-016", "Implementar autenticación OAuth2", "AuthService", "Google + GitHub providers"),
        ("TK-017", "Setup monitoring con Prometheus", "DevOps", "Alertas por Slack"),
        ("TK-018", "Containerizar aplicación con Docker", "DevOps", "Multi-stage build"),
        ("TK-019", "Implementar backup automático", "Infrastructure", "S3 bucket + cron"),
        ("TK-020", "Revisar vulnerabilidades de seguridad", "Security", "OWASP Top 10"),
        ("TK-021", "Optimizar performance del servidor", "Infrastructure", ""),
    ],
    "María López": [
        ("TK-022", "Testing E2E con Playwright", "QA", "Escenarios críticos de negocio"),
        ("TK-023", "Unit tests módulo de facturación", "BillingModule", "Cobertura > 80%"),
        ("TK-024", "Documentar casos de prueba", "QA", "Usar formato Gherkin"),
        ("TK-025", "Validación de formularios frontend", "FrontendApp", "Zod schema validation"),
        ("TK-026", "Performance testing con k6", "QA", "500 usuarios concurrentes"),
    ],
    "Roberto Silva": [
        ("TK-027", "API de notificaciones push", "NotificationService", "Firebase FCM"),
        ("TK-028", "Módulo de exportación PDF", "ReportsModule", "jsPDF + autoTable"),
        ("TK-029", "Integración con Stripe para pagos", "PaymentGateway", "Webhooks incluidos"),
        ("TK-030", "Dashboard en tiempo real con WebSockets", "Analytics", "Socket.io"),
        ("TK-031", "Sistema de roles y permisos", "AuthService", "RBAC pattern"),
        ("TK-032", "Internacionalización i18n", "FrontendApp", "ES, EN, PT"),
        ("TK-033", "Microservicio de emails", "NotificationService", "SendGrid API"),
        ("TK-034", "Caché a nivel de aplicación", "Infrastructure", ""),
        ("TK-035", "Refactorizar módulo legacy", "BackgroundJobs", "Código de 2019"),
    ],
    "Patricia Romero": [
        ("TK-036", "Análisis de requerimientos módulo CRM", "CRMModule", "Reunión con stakeholders"),
        ("TK-037", "Implementar CRUD de contactos", "CRMModule", ""),
        ("TK-038", "Pipeline de datos con Apache Kafka", "DataPipeline", "Topics definidos en Confluence"),
        ("TK-039", "Integración con Salesforce", "CRMModule", "API REST v55.0"),
        ("TK-040", "Implementar búsqueda full-text", "SearchService", "Elasticsearch 8.x"),
    ],
}

def create_sample_excel(output_path):
    wb = Workbook()
    # Eliminar hoja por defecto
    wb.remove(wb.active)

    # Colores de cabecera
    HEADER_FILL  = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    HEADER_FONT  = Font(color="FFFFFF", bold=True, size=11)
    BORDER_SIDE  = Side(style="thin", color="D1D5DB")
    CELL_BORDER  = Border(left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE)

    for prog_name, tickets in PROGRAMMERS.items():
        ws = wb.create_sheet(title=prog_name)

        # Encabezados
        headers = ["N° Ticket", "Descripcion", "Proyecto", "Notas"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill  = HEADER_FILL
            cell.font  = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = CELL_BORDER

        ws.row_dimensions[1].height = 22

        # Datos
        ALT_FILL = PatternFill(start_color="F8F7FF", end_color="F8F7FF", fill_type="solid")
        for row_idx, ticket in enumerate(tickets, 2):
            for col_idx, value in enumerate(ticket, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = CELL_BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=(col_idx in [2, 4]))
                if row_idx % 2 == 0:
                    cell.fill = ALT_FILL

        # Ajustar anchos de columna
        col_widths = [12, 50, 22, 35]
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(output_path)
    print(f"✓ Archivo Excel generado: {output_path}")
    print(f"  Programadores: {len(PROGRAMMERS)}")
    total = sum(len(t) for t in PROGRAMMERS.values())
    print(f"  Total de tickets: {total}")

if __name__ == "__main__":
    script_dir = os.path.dirname(os.path.abspath(__file__))
    out = os.path.join(script_dir, "ejemplo_tickets.xlsx")
    create_sample_excel(out)
